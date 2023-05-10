from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook("example.xlsx")
ws = wb.active
coursea = ws["A1"].value
courseb = ws["B1"].value
coursec = ws["C1"].value
print("A코너 : ", coursea)
print("B코너 : ", courseb)
print("C코너 : ", coursec)