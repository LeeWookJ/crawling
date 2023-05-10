from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from openpyxl import Workbook
# import psycopg2

driver = webdriver.Chrome()
driver.get("https://www.inje.ac.kr/kor/Template/Bsub_page.asp?Ltype=5&Ltype2=3&Ltype3=3&Tname=S_Food&Ldir=board/S_Food&Lpage=s_food_view&d1n=5&d2n=4&d3n=4&d4n=0")
PATH_CAFETERIA_DAY = r'//*[@id="table1"]/thead/tr/th[5]'

PATH_CAFETERIA_SECTION_A = r'//*[@id="table1"]/tbody/tr[1]/td[1]'
PATH_CAFETERIA_SECTION_B= r'//*[@id="table1"]/tbody/tr[2]/td[1]'
PATH_CAFETERIA_SECTION_C = r'//*[@id="table1"]/tbody/tr[3]/td[1]'

PATH_CAFETERIA_COURSE_A = r'//*[@id="table1"]/tbody/tr[1]/td[5]'
PATH_CAFETERIA_COURSE_B = r'//*[@id="table1"]/tbody/tr[2]/td[5]'
PATH_CAFETERIA_COURSE_C = r'//*[@id="table1"]/tbody/tr[3]/td[5]'

def crawling(path):
    element = driver.find_element(By.XPATH, path)

    if element == None:
        raise Exception("찾으려는 원소가 없습니다.")
    return element.text
schoolcafeteriaday = crawling(PATH_CAFETERIA_DAY)

schoolcafeteriasectiona = crawling(PATH_CAFETERIA_SECTION_A)
schoolcafeteriasectionb = crawling(PATH_CAFETERIA_SECTION_B)
schoolcafeteriasectionc = crawling(PATH_CAFETERIA_SECTION_C)

schoolcafeteriacoursea = crawling(PATH_CAFETERIA_COURSE_A)
schoolcafeteriacourseb = crawling(PATH_CAFETERIA_COURSE_B)
schoolcafeteriacoursec = crawling(PATH_CAFETERIA_COURSE_C)
wb = Workbook()
ws = wb.active
ws.title = "example Sheet"
ws["A1"].value = schoolcafeteriacoursea
ws["B1"].value = schoolcafeteriacourseb
ws["C1"].value = schoolcafeteriacoursec
wb.save("example.xlsx")
print(schoolcafeteriaday+'\n\n'
      ,schoolcafeteriasectiona,schoolcafeteriacoursea+'\n\n'
      ,schoolcafeteriasectionb,schoolcafeteriacourseb+'\n\n'
      ,schoolcafeteriasectionc,schoolcafeteriacoursec)
# 코드개선점 : 변수명칭, 코드가 난잡함, class써보기, 클린코드
# 기능개선점 : 여러 날짜의값 가져오기, 주마다 갱신하기
# db = psycopg2.connect(host={서버 포트}, dbname={데이터베이스 명},user={서버의 데이터베이스 유저명},password={해당 유저 비밀번호},port=5432)
#assert "Python" in driver.title
#elem = driver.find_element(By.NAME, "q")
#elem.clear()
#elem.send_keys("pycon")
#elem.send_keys(Keys.RETURN)
#assert "No results found." not in driver.page_source
#driver.close()